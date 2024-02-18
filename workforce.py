import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill
import sys

# Function to read input DataFrame from Excel file
def read_input_data(input_file_path):
    df_input = pd.read_excel(input_file_path)
    # Additional preprocessing steps if needed
    return df_input

# Function to preprocess input DataFrame
def preprocess_input_data(df_input):
   # Convert 'Date' column to datetime format
    df_input['Date'] = pd.to_datetime(df_input['Date'], format='%d-%m-%Y')

    # Extract start and end times from 'timing' column and remove extra spaces
    df_input[['Start Time', 'End Time']] = df_input['timing'].str.split('-', expand=True)
    df_input['Start Time'] = pd.to_datetime(df_input['Start Time'].str.strip(), format='%H:%M')
    df_input['End Time'] = pd.to_datetime(df_input['End Time'].str.strip(), format='%H:%M')

    # Get the timestamp for the start and end times
    start_timestamp = pd.to_datetime(df_input['Date'].iloc[0].replace(hour=df_input['Start Time'].iloc[0].hour, minute=df_input['Start Time'].iloc[0].minute))
    end_timestamp = pd.to_datetime(df_input['Date'].iloc[0].replace(hour=df_input['End Time'].iloc[0].hour, minute=df_input['End Time'].iloc[0].minute))

    print('start time', start_timestamp)
    print('end Time', end_timestamp)
    return start_timestamp,end_timestamp

# Function to filter data based on time range
def filter_data(MainData_df, start_timestamp, end_timestamp):
   # Check if 'Actual Arrival' column exists in MainData_df
    if 'Actual Arrival' in MainData_df.columns:
        # Use 'Actual Arrival' if not null, otherwise fallback to 'Sched Arrival'
        MainData_df['Arrival'] = MainData_df['Actual Arrival'].combine_first(MainData_df['Sched Arrival'])
        Actual_Arrival = MainData_df['Arrival']
        
    else:
        # Use 'Sched Arrival' if 'Actual Arrival' column doesn't exist
        MainData_df['Arrival'] = MainData_df['Sched Arrival']
        Actual_Arrival = MainData_df['Arrival']

    MainData_df['Arrival'] = pd.to_datetime(MainData_df['Arrival'], format='%Y-%m-%d %H:%M:%S', errors='coerce')
    
    # Filter the original DataFrame based on the input date and time range
    filtered_data = MainData_df[
        (MainData_df['Arrival'] >= start_timestamp)
        & (MainData_df['Arrival'] <= end_timestamp)
    ].copy()
    return filtered_data

# Function to calculate workforce for a given time interval
def calculate_workforce(filtered_data, Total_Minutes_calculation):

    row_count = len(filtered_data)
    filtered_data['Dumper'] = filtered_data['EXTRA SMALL'] + filtered_data['SMALL']
    filtered_data['Infeed'] = filtered_data['MEDIUM'] + filtered_data['LARGE']
    filtered_data['LL'] = filtered_data['EXTRA LARGE'] + filtered_data['NC'] + filtered_data['NC PLUS'] + filtered_data['HEAVY BULKY'] + filtered_data['HEAVY BULKY PLUS']
    filtered_data['XD'] = filtered_data['Xdock Packages']

    total_dumper = filtered_data['Dumper'].sum()
    total_infeed = filtered_data['Infeed'].sum()
    total_sortable = total_dumper + total_infeed
    total_LL = filtered_data['LL'].sum()
    total_XD = filtered_data['XD'].sum()
    total_Volume = total_XD + total_dumper + total_infeed + total_LL
    row_count = len(filtered_data)
    total_unloader = round((row_count * 45) / Total_Minutes_calculation)
    total_injectors = round(total_infeed / ((700 / 60) * (Total_Minutes_calculation)))
    total_facers = round(total_dumper / ((2300 / 60) * (Total_Minutes_calculation)))
    if total_dumper >= 9000:
        total_dumper_operators = 2
    elif total_dumper == 0:
        total_dumper_operators = 0
    else:
        total_dumper_operators = 1
    
    

    #Printing Selected columns
    selected_columns = filtered_data[['Route','Load ID','Sched Arrival','Actual Arrival','Dumper', 'Infeed', 'LL']]

    return total_unloader, selected_columns,total_injectors, total_facers, total_dumper_operators, total_Volume,total_dumper,total_infeed,total_sortable,total_LL,total_XD,row_count

# Function to write output to Excel file
def write_output_to_excel(output_sheet, selected_columns, total_dumper, total_infeed, total_sortable, total_LL, total_XD, row_count, total_unloader, total_injectors, total_facers, total_dumper_operators, total_Volume):
    # Write column names to the 'Output' sheet starting from cell P6
    for idx, col in enumerate(selected_columns.columns):
        col_letter = chr(ord('I') + idx)  # P, Q, R, ...
        output_sheet[col_letter + '1'] = col

    # Write selected_columns values to the 'Output' sheet starting from cell P7
    for idx, col in enumerate(selected_columns.columns):
        col_letter = chr(ord('I') + idx)  # P, Q, R, ...
        start_row = 2
        for i, value in enumerate(selected_columns[col]):
            output_sheet[col_letter + str(start_row + i)] = value

    # Set font style for labels
    font_labels = Font(bold=True, color='000000', size=12)
    for label_cell in ['E3', 'E4', 'E5', 'E6', 'E7', 'A1', 'A3', 'A4', 'A5', 'A6', 'A7', 'A9', 'E9', 'E10', 'E1']:
        output_sheet[label_cell].font = font_labels

    # Set font style for values
    font_values = Font(color='000000', size=12)
    for value_cell in ['F3', 'F4', 'F5', 'F6', 'F7', 'B1', 'F1', 'B3', 'B4', 'B5', 'B7']:
        output_sheet[value_cell].font = font_values

    # Set background color for headers
    fill_headers = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
    for header_cell in ['E3', 'E4', 'E5', 'E6', 'E7', 'A1', 'A3', 'A4', 'A5', 'A6', 'A7', 'A9', 'E9', 'E10', 'E1']:
        output_sheet[header_cell].fill = fill_headers

    # Set background color for values
    fill_values = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for value_cell in ['F3', 'F4', 'F5', 'F6', 'F7', 'B1', 'F1', 'B3', 'B4', 'B5', 'B7']:
        output_sheet[value_cell].fill = fill_values

    # Write the labels to cells M7 and M8
    output_sheet['E3'] = 'Dumper'
    output_sheet['E4'] = 'Infeed'
    output_sheet['E5'] = 'Sortable'
    output_sheet['E6'] = 'LL'
    output_sheet['E7'] = 'XD'
    output_sheet['A1'] = 'Total_Volume'
    output_sheet['A3'] = 'Unloaders'
    output_sheet['A4'] = 'Injectors'
    output_sheet['A5'] = 'Facers'
    output_sheet['A6'] = 'LL'
    output_sheet['A7'] = 'Dumper Operators'
    output_sheet['A9'] = 'Required Speed'
    output_sheet['E9'] = 'Required TPH'
    output_sheet['E10'] = 'Actual TPH'
    output_sheet['E1'] = 'No of Trucks'
    
    # Write the total sums to cells N7 and N8
    output_sheet['F3'] = total_dumper
    output_sheet['F4'] = total_infeed
    output_sheet['F5'] = total_sortable
    output_sheet['F6'] = total_LL
    output_sheet['F7'] = total_XD
    output_sheet['B1'] = total_Volume
    output_sheet['F1'] = row_count
    output_sheet['B3'] = total_unloader
    output_sheet['B4'] = total_injectors
    output_sheet['B5'] = total_facers
    output_sheet['B7'] = total_dumper_operators
    return output_sheet

# Function to calculate workforce for the whole day
def calculate_workforce_whole_day(MainData_df,workbook):
    
    whole_day_sheet = workbook['WholeDay']

    for hour in range(24):
        # Calculate the start and end hours for the current interval
        start_hour = hour
        end_hour = (hour + 1) % 24  # Wrap around to 0 if hour + 1 exceeds 23
        
        # Construct the time interval string
        time_interval = f"{start_hour:02}:00 - {end_hour:02}:00"
        
        # Convert the time interval to datetime objects
        start_timestamp_wholeDay = pd.to_datetime(f"2023-12-22 {start_hour:02}:00:00")
        end_timestamp_wholeDay = pd.to_datetime(f"2023-12-22 {end_hour:02}:00:00")
        
        # Filter the data for the current time interval
        filtered_data_wholeDay = MainData_df[
            (MainData_df['Arrival'] >= start_timestamp_wholeDay)
            & (MainData_df['Arrival'] < end_timestamp_wholeDay)
        ].copy()
        
        # Print or process the filtered data for the current time interval

        row_count = len(filtered_data_wholeDay)
        filtered_data_wholeDay['Dumper'] = filtered_data_wholeDay['EXTRA SMALL'] + filtered_data_wholeDay['SMALL']
        filtered_data_wholeDay['Infeed'] = filtered_data_wholeDay['MEDIUM'] + filtered_data_wholeDay['LARGE']
        filtered_data_wholeDay['LL'] = filtered_data_wholeDay['EXTRA LARGE'] + filtered_data_wholeDay['NC'] + filtered_data_wholeDay['NC PLUS'] + filtered_data_wholeDay['HEAVY BULKY'] + filtered_data_wholeDay['HEAVY BULKY PLUS']
        filtered_data_wholeDay['XD'] = filtered_data_wholeDay['Xdock Packages']

        Total_Minutes_calculation_wholeDay  = 60
        # Calculate the total sum of the 'Dumper' and 'Infeed' columns
        total_dumper = filtered_data_wholeDay['Dumper'].sum()
        total_infeed = filtered_data_wholeDay['Infeed'].sum()
        total_sortable = total_dumper + total_infeed
        total_LL = filtered_data_wholeDay['LL'].sum()
        total_XD = filtered_data_wholeDay['XD'].sum()
        total_Volume = total_XD + total_dumper + total_infeed + total_LL
        total_unloader = round((row_count * 45) / Total_Minutes_calculation_wholeDay)
        total_injectors = round(total_infeed / ((700 / 60) * (Total_Minutes_calculation_wholeDay)))
        total_facers = round(total_dumper / ((2300 / 60) * (Total_Minutes_calculation_wholeDay)))
    
        #total_dumper_operators = 2 if total_dumper >= 9000 else if 1
        if(total_dumper>=9000):
            total_dumper_operators = 2
        elif(total_dumper==0):
            total_dumper_operators = 0
        else:
            total_dumper_operators = 1

        # Write the time interval to the corresponding cell in the 'WholeDay' sheet
        whole_day_sheet.cell(row=1, column=hour + 2, value=time_interval)
         # Write totals to the corresponding cells in the 'WholeDay' sheet
        whole_day_sheet.cell(row=2, column=hour + 2, value=total_dumper)
        whole_day_sheet.cell(row=3, column=hour + 2, value=total_infeed)
        whole_day_sheet.cell(row=4, column=hour + 2, value=total_sortable)
        whole_day_sheet.cell(row=5, column=hour + 2, value=total_LL)
        whole_day_sheet.cell(row=6, column=hour + 2, value=total_XD)
        whole_day_sheet.cell(row=7, column=hour + 2, value=total_unloader)
        whole_day_sheet.cell(row=8, column=hour + 2, value=total_injectors)
        whole_day_sheet.cell(row=9, column=hour + 2, value=total_facers)
        whole_day_sheet.cell(row=11, column=hour + 2, value=total_dumper_operators)
        whole_day_sheet.cell(row=13, column=hour + 2, value=total_Volume)
    return whole_day_sheet

def main():
    
    input_file_path = r"C:\Users\Legion 5pro\Downloads\Input.xlsx"
    output_file_path = r"c:\Users\Legion 5pro\Downloads\SingleSheet.xlsx"
    workbook = openpyxl.load_workbook(output_file_path)
    try:
        # Read input DataFrame
        df_input = read_input_data(input_file_path)
        # Preprocess input DataFrame
        start_timestamp,end_timestamp = preprocess_input_data(df_input)

        # Define start and end timestamps
        # Assuming start and end timestamps are provided or calculated
        # Load the workbook
        workbook = openpyxl.load_workbook(output_file_path)

        # Get the 'MainData' sheet
        MainData_sheet = workbook['MainData']
        MainData_df = pd.DataFrame(MainData_sheet.values, columns=[col[0].value for col in MainData_sheet.iter_cols()])
        # Filter data based on time range
        filtered_data = filter_data(MainData_df, start_timestamp, end_timestamp)

        print("filtered_data : ",filtered_data)
        # Calculate Total_Minutes_calculation (time difference in minutes between start and end timestamps)
        Total_Minutes_calculation  = (end_timestamp - start_timestamp).total_seconds() / 60

        # Calculate workforce for the specified time range
        total_unloader, selected_columns,total_injectors, total_facers, total_dumper_operators, total_Volume,total_dumper, total_infeed, total_sortable, total_LL, total_XD, row_count = calculate_workforce(filtered_data, Total_Minutes_calculation)

        # Create or get the 'Output' sheet
        workbook = openpyxl.load_workbook(output_file_path)
        output_sheet = workbook['Output']


        # Write output to Excel file
        write_output_to_excel(output_sheet, selected_columns, total_dumper, total_infeed, total_sortable, total_LL, total_XD, row_count, total_unloader, total_injectors, total_facers, total_dumper_operators, total_Volume)

        # Calculate workforce for the whole day
        workforce_data_whole_day = calculate_workforce_whole_day(MainData_df,workbook)

        # Save changes to the Excel file
        workbook.save(output_file_path)

    except FileNotFoundError as e:
        print(f"File not found: {e}")
    except PermissionError as e:
        print(f"Permission denied: {e}")
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    main()
